#/usr/bin/env python3
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def data_value(data):
    data_list = []
    for line in data:
        line_list =[]
        for i in line:
            line_list.append(i.value)
        data_list.append(line_list)
    return data_list

def Data_merging(students_list,time_list):
    data = []
    for i in students_list:
        x, y, z = i
        for j in time_list:
            o, p, t = j
            if y == p :
                line_data = x,y,z,t
                data.append(line_data)
    return data



def combine():
    xlsx = load_workbook('courses.xlsx')
    students_tab = xlsx['students']
    time_tab = xlsx['time']
    students_list = data_value(students_tab)
    time_list = data_value(time_tab)
    data = Data_merging(students_list,time_list)
    ws = xlsx.create_sheet('combine')
    for row in range(len(data)):    
        ws.append(data[row])
    xlsx.save('courses.xlsx')


'''
将数据按年分类
'''
def data_dict(combine_list):
    s = set()
    d = {}
    for i in range(1,len(combine_list)):
        #print(combine_list[i][0].strftime('%Y'))
        s.add(combine_list[i][0].strftime('%Y'))
        
        for j in s :
            if combine_list[i][0].strftime('%Y') == j:
                
                if d.get(j):
                    d[j] = d.get(j) + [combine_list[i]]
                else:
                    d[j] = [combine_list[i]]
    return d


def split():
    xlsx = load_workbook('courses.xlsx',read_only=True)
    combine_list = data_value(xlsx['combine'])
    combine_dict = data_dict(combine_list)
    title1,title2,title3,title4 = combine_list[0]
    for title,date in combine_dict.items():
        wb = Workbook()
        ws = wb.active
        ws.title = title        
        ws.append([title1, title2, title3, title4])
        for row in range(len(date)):
            ws.append(date[row])
        wb.save(title+'.xlsx')




if __name__ == '__main__':
    combine()
    split()
